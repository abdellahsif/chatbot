from __future__ import annotations

import csv
import json
import os
import re
import unicodedata
from pathlib import Path
from statistics import mean
from typing import Any

from app.supabase_store import fetch_schools


class DataBundle:
    def __init__(self, schools: dict[str, dict], transcripts: list[dict], policy: dict, source: str = "unknown"):
        self.schools = schools
        self.transcripts = transcripts
        self.policy = policy
        self.source = source


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: Any) -> bool:
    text = _safe_str(value).lower()
    return text in {"true", "1", "yes", "oui", "y", "vrai"}


def _parse_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = _safe_str(value)
    if not text:
        return default
    text = text.replace(" ", "").replace(",", ".")
    try:
        return int(float(text))
    except ValueError:
        return default


def _to_iso_date(value: Any) -> str:
    if value is None:
        return "2026-01-01"
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = _safe_str(value)
    if not text:
        return "2026-01-01"
    return text


def _normalize_token(value: Any) -> str:
    text = _safe_str(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "general"


def _split_program_values(value: Any) -> list[str]:
    text = _safe_str(value)
    if not text:
        return []
    parts = re.split(r"[|,;/]+", text)
    out: list[str] = []
    for part in parts:
        token = _normalize_token(part)
        if token and token not in out:
            out.append(token)
    return out


def _split_program_labels(value: Any) -> list[str]:
    text = _safe_str(value)
    if not text:
        return []
    parts = re.split(r"[|,;/]+", text)
    labels: list[str] = []
    for part in parts:
        clean = " ".join(_safe_str(part).split())
        if clean and clean not in labels:
            labels.append(clean)
    return labels


def _detect_language(text: str) -> str:
    if not text:
        return "fr"
    text_lower = _safe_str(text).lower()
    if re.search(r"[\u0600-\u06FF]", text_lower):
        return "ar"
    return "fr"


def _preserve_accents_text(text: str) -> str:
    return text


def _build_metadata(
    *,
    school_name: str,
    city: str,
    status: str,
    program: str,
    domain: str,
    tuition_min: int | None,
    tuition_max: int | None,
    admission_type: str,
    source_url: str,
) -> dict[str, Any]:
    return {
        "school_name": school_name,
        "city": city,
        "status": status,
        "program": program,
        "domain": domain,
        "tuition_min": tuition_min,
        "tuition_max": tuition_max,
        "admission_type": admission_type,
        "source": "official" if source_url else "unknown",
    }


def _build_chunk(
    *,
    chunk_id: str,
    school_id: str,
    chunk_type: str,
    text: str,
    language: str,
    metadata: dict[str, Any],
    created_at: str,
    source_url: str,
    source_confidence: float,
    program: str,
) -> dict[str, Any]:
    return {
        "chunk_id": chunk_id,
        "school_id": school_id,
        "chunk_type": chunk_type,
        "language": language,
        "metadata": metadata,
        "text": _preserve_accents_text(text),
        "chunk_version": "v2",
        "created_at": created_at,
        "source_confidence": source_confidence,
        "source_url": source_url,
        "program": program,
        "recorded_at": created_at,
    }


def _infer_school_type(row: dict[str, Any]) -> str:
    status = _safe_str(row.get("legal_status") or row.get("status") or row.get("type")).lower()
    if "public" in status:
        return "public"
    if "prive" in status or "priv" in status or "private" in status:
        return "private"
    return "public"


def _infer_tuition_bounds(row: dict[str, Any], school_type: str) -> tuple[int, int]:
    pricing_min = _parse_int(row.get("pricing_min"), default=-1)
    pricing_max = _parse_int(row.get("pricing_max"), default=-1)
    if pricing_min < 0:
        pricing_min = _parse_int(row.get("tuition_min_mad"), default=-1)
    if pricing_max < 0:
        pricing_max = _parse_int(row.get("tuition_max_mad"), default=-1)

    if pricing_min < 0 and pricing_max < 0:
        return (-1, -1)
    if pricing_min < 0:
        pricing_min = max(0, pricing_max)
    if pricing_max < 0:
        pricing_max = max(0, pricing_min)
    if pricing_min == 0 and pricing_max == 0 and school_type != "public":
        return (-1, -1)
    return pricing_min, pricing_max


def load_from_supabase_schools(limit: int = 500) -> tuple[dict[str, dict], list[dict]]:
    response = fetch_schools(limit=limit)
    rows = response.get("items", []) if isinstance(response, dict) else []
    if not isinstance(rows, list) or not rows:
        return {}, []

    schools: dict[str, dict] = {}
    transcripts: list[dict] = []

    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue

        raw_id = _safe_str(row.get("id"))
        if not raw_id:
            continue

        name = _safe_str(row.get("name")) or f"school_{idx}"
        city = _safe_str(row.get("city"))
        school_id = raw_id
        school_type = _infer_school_type(row)
        pricing_min, pricing_max = _infer_tuition_bounds(row, school_type)

        programs = _split_program_values(row.get("programs_tags"))
        filieres = _split_program_values(row.get("filieres"))
        all_programs = sorted(set(programs + filieres)) or ["general"]
        legal_status = _safe_str(row.get("legal_status"))
        conditions = _safe_str(row.get("conditions"))
        domaine_principal = _safe_str(row.get("domaine_principal"))

        pretty_programs = _split_program_labels(row.get("programs_tags"))
        pretty_filieres = _split_program_labels(row.get("filieres"))
        pretty_program_labels = pretty_programs + [x for x in pretty_filieres if x not in pretty_programs]

        schools[school_id] = {
            "school_id": school_id,
            "name": name,
            "country": "MA",
            "city": city,
            "type": school_type,
            "legal_status": legal_status,
            "conditions": conditions,
            "tuition_min_mad": pricing_min if pricing_min >= 0 else 0,
            "tuition_max_mad": pricing_max if pricing_max >= 0 else 0,
            "pricing_min": pricing_min if pricing_min >= 0 else 0,
            "pricing_max": pricing_max if pricing_max >= 0 else 0,
            "pricing_details": row.get("pricing_details"),
            "programs": all_programs,
            "programs_tags": _safe_str(row.get("programs_tags")),
            "filieres": _safe_str(row.get("filieres")),
            "domaine_principal": domaine_principal,
            "admission_selectivity": "medium",
            "employability_score": 3.8,
            "salary_entry_min_mad": 6000,
            "salary_entry_max_mad": 12000,
            "international_double_degree": False,
            "website_url": _safe_str(row.get("website_url")),
            "logo_url": _safe_str(row.get("logo_url")),
            "acronym": _safe_str(row.get("acronym")),
            "source": "supabase_schools",
            "source_row_id": raw_id,
        }

        created_at = _to_iso_date(row.get("created_at"))
        source_url = _safe_str(row.get("website_url"))
        source_confidence = 0.9 if source_url else 0.7
        tuition_min_verified = pricing_min if pricing_min >= 0 else None
        tuition_max_verified = pricing_max if pricing_max >= 0 else None
        status = legal_status or school_type

        detected_lang = _detect_language(name + " " + city + " " + status)
        display_lang = "ar" if detected_lang == "ar" else "fr"

        overview_parts = [
            f"{name} est un établissement {status} situé à {city}." if city else f"{name} est un établissement {status}."
        ]
        if domaine_principal:
            overview_parts.append(f"Domaines principaux : {domaine_principal}.")
        overview_text = " ".join(p for p in overview_parts if p).strip()

        base_metadata = _build_metadata(
            school_name=name,
            city=city,
            status=status,
            program="",
            domain=domaine_principal,
            tuition_min=tuition_min_verified,
            tuition_max=tuition_max_verified,
            admission_type=conditions,
            source_url=source_url,
        )

        transcripts.append(
            _build_chunk(
                chunk_id=f"{_normalize_token(raw_id)}_overview",
                school_id=school_id,
                chunk_type="overview",
                text=overview_text,
                language=display_lang,
                metadata=base_metadata,
                created_at=created_at,
                source_url=source_url,
                source_confidence=source_confidence,
                program=all_programs[0] if all_programs else "general",
            )
        )

        if domaine_principal:
            domain_text = f"Domaine principal : {domaine_principal}."
            domain_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domaine_principal,
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{_normalize_token(raw_id)}_domain_{_normalize_token(domaine_principal)}",
                    school_id=school_id,
                    chunk_type="domain",
                    text=domain_text,
                    language=display_lang,
                    metadata=domain_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=all_programs[0] if all_programs else "general",
                )
            )

        program_labels = pretty_program_labels if pretty_program_labels else [p for p in all_programs if _safe_str(p)]
        for prog_label in program_labels:
            clean_label = _safe_str(prog_label)
            if not clean_label:
                continue
            prog_token = _normalize_token(clean_label)
            program_text = f"{name} propose la formation {clean_label}."
            program_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program=clean_label,
                domain=domaine_principal,
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{_normalize_token(raw_id)}_program_{prog_token}",
                    school_id=school_id,
                    chunk_type="program",
                    text=program_text,
                    language=display_lang,
                    metadata=program_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=prog_token,
                )
            )

        if tuition_min_verified is not None or tuition_max_verified is not None:
            if tuition_min_verified is not None and tuition_max_verified is not None:
                tuition_text = f"Les frais de scolarité varient entre {tuition_min_verified} et {tuition_max_verified} MAD par an."
            else:
                value = tuition_min_verified if tuition_min_verified is not None else tuition_max_verified
                tuition_text = f"Les frais de scolarité sont estimés à {value} MAD par an."
            tuition_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domaine_principal,
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{_normalize_token(raw_id)}_tuition",
                    school_id=school_id,
                    chunk_type="tuition",
                    text=tuition_text,
                    language=display_lang,
                    metadata=tuition_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=all_programs[0] if all_programs else "general",
                )
            )

        if conditions:
            admission_text = f"Admission : {conditions}."
            admission_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domaine_principal,
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type=conditions,
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{_normalize_token(raw_id)}_admission",
                    school_id=school_id,
                    chunk_type="admission",
                    text=admission_text,
                    language=display_lang,
                    metadata=admission_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=all_programs[0] if all_programs else "general",
                )
            )

    return schools, transcripts


def _sheet_rows_by_headers(xlsx_path: Path) -> dict[str, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    tables: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [_safe_str(h) for h in header_row]
        key = "|".join(headers)
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            if any(v is not None and _safe_str(v) for v in item.values()):
                rows.append(item)
        if rows:
            tables[key] = rows
    return tables


def _load_from_excel_catalog_tables(tables: dict[str, list[dict[str, Any]]]) -> tuple[dict[str, dict], list[dict]]:
    primary_rows: list[dict[str, Any]] = []
    for rows in tables.values():
        if not rows:
            continue
        sample = rows[0]
        if "Nom" in sample and "Ville" in sample and ("Statut" in sample or "Type" in sample):
            primary_rows = rows
            break

    schools: dict[str, dict] = {}
    transcripts: list[dict] = []
    if not primary_rows:
        return schools, transcripts

    for i, row in enumerate(primary_rows, start=1):
        name = _safe_str(row.get("Nom"))
        if not name:
            continue

        city = _safe_str(row.get("Ville"))
        status = _safe_str(row.get("Statut"))
        school_type = "public" if "public" in status.lower() else "private"
        domain = _safe_str(row.get("Domaine"))
        filieres_text = _safe_str(row.get("Filières"))
        programs = _split_program_values(filieres_text or domain)
        if not programs:
            programs = ["general"]

        tuition_min = _parse_int(row.get("Frais inscription (MAD)"), default=0)
        tuition_max = _parse_int(row.get("Frais annuels (MAD)"), default=0)
        if tuition_max <= 0 and school_type == "public":
            tuition_max = max(tuition_min, 1200)

        conditions = _safe_str(row.get("Conditions d'accès"))
        selectivity = "high" if any(k in conditions.lower() for k in ["concours", "selectif", "sélectif"]) else "medium"

        school_id = f"xl_{_normalize_token(name)}_{_normalize_token(city)}_{i}"
        schools[school_id] = {
            "school_id": school_id,
            "name": name,
            "country": "MA",
            "city": city,
            "type": school_type,
            "legal_status": status,
            "conditions": conditions,
            "tuition_min_mad": tuition_min,
            "tuition_max_mad": tuition_max,
            "pricing_min": tuition_min,
            "pricing_max": tuition_max,
            "pricing_details": _safe_str(row.get("Frais par filière (détail)")),
            "programs": programs,
            "programs_tags": filieres_text,
            "filieres": filieres_text,
            "admission_selectivity": selectivity,
            "employability_score": 3.7,
            "salary_entry_min_mad": 5000,
            "salary_entry_max_mad": 12000,
            "international_double_degree": False,
            "website_url": _safe_str(row.get("Site Web")),
            "logo_url": "",
            "acronym": _safe_str(row.get("Sigle")),
            "source": "excel_catalog",
            "source_row_id": str(i),
        }

        detected_lang = _detect_language(name + " " + city + " " + domain)
        display_lang = "ar" if detected_lang == "ar" else "fr"
        created_at = _to_iso_date(row.get("Date de collecte"))
        source_url = _safe_str(row.get("Site Web"))
        source_confidence = 0.85 if source_url else 0.7
        status = status or school_type

        base_metadata = _build_metadata(
            school_name=name,
            city=city,
            status=status,
            program="",
            domain=domain,
            tuition_min=tuition_min if tuition_min > 0 else None,
            tuition_max=tuition_max if tuition_max > 0 else None,
            admission_type=conditions,
            source_url=source_url,
        )

        overview_text = (
            f"{name} est un établissement {status} situé à {city}." if city else f"{name} est un établissement {status}."
        )
        if domain:
            overview_text += f" Domaines principaux : {domain}."

        transcripts.append(
            _build_chunk(
                chunk_id=f"{school_id}_overview",
                school_id=school_id,
                chunk_type="overview",
                text=overview_text,
                language=display_lang,
                metadata=base_metadata,
                created_at=created_at,
                source_url=source_url,
                source_confidence=source_confidence,
                program=programs[0],
            )
        )

        if domain:
            domain_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_domain_{_normalize_token(domain)}",
                    school_id=school_id,
                    chunk_type="domain",
                    text=f"Domaine principal : {domain}.",
                    language=display_lang,
                    metadata=domain_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0],
                )
            )

        for label in _split_program_labels(filieres_text):
            prog_label = _safe_str(label)
            if not prog_label:
                continue
            prog_token = _normalize_token(prog_label)
            program_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program=prog_label,
                domain=domain,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_program_{prog_token}",
                    school_id=school_id,
                    chunk_type="program",
                    text=f"{name} propose la formation {prog_label}.",
                    language=display_lang,
                    metadata=program_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=prog_token,
                )
            )

        if tuition_min > 0 or tuition_max > 0:
            tuition_text = f"Les frais de scolarité varient entre {tuition_min} et {tuition_max} MAD par an."
            tuition_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain,
                tuition_min=tuition_min,
                tuition_max=tuition_max,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_tuition",
                    school_id=school_id,
                    chunk_type="tuition",
                    text=tuition_text,
                    language=display_lang,
                    metadata=tuition_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0],
                )
            )

        if conditions:
            admission_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type=conditions,
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_admission",
                    school_id=school_id,
                    chunk_type="admission",
                    text=f"Admission : {conditions}.",
                    language=display_lang,
                    metadata=admission_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0],
                )
            )

    return schools, transcripts


def load_from_excel_mcd(xlsx_path: Path) -> tuple[dict[str, dict], list[dict]]:
    tables = _sheet_rows_by_headers(xlsx_path)

    catalog_schools, catalog_transcripts = _load_from_excel_catalog_tables(tables)
    if catalog_schools:
        return catalog_schools, catalog_transcripts

    etab: list[dict[str, Any]] = []
    filiere: list[dict[str, Any]] = []
    cout: list[dict[str, Any]] = []
    candidature: list[dict[str, Any]] = []
    metier: list[dict[str, Any]] = []
    mener: list[dict[str, Any]] = []

    for key, rows in tables.items():
        cols = set(key.split("|"))
        if {"id_etablissement", "nom_etablissement"}.issubset(cols):
            etab = rows
        elif {"id_filiere", "id_etablissement", "nom_filiere"}.issubset(cols):
            filiere = rows
        elif {"id_cout", "id_filiere", "frais_scolarite_annuel"}.issubset(cols):
            cout = rows
        elif {"id_candidature", "id_filiere", "procedure"}.issubset(cols):
            candidature = rows
        elif {"id_metier", "nom_metier", "salaire_debutant_maroc"}.issubset(cols):
            metier = rows
        elif {"id_filiere", "id_metier"}.issubset(cols):
            mener = rows

    etab_by_id = {_safe_str(r.get("id_etablissement")): r for r in etab}
    filiere_by_id = {_safe_str(r.get("id_filiere")): r for r in filiere}

    filieres_by_etab: dict[str, list[dict[str, Any]]] = {}
    for row in filiere:
        etab_id = _safe_str(row.get("id_etablissement"))
        filieres_by_etab.setdefault(etab_id, []).append(row)

    cout_by_filiere: dict[str, list[dict[str, Any]]] = {}
    for row in cout:
        filiere_id = _safe_str(row.get("id_filiere"))
        cout_by_filiere.setdefault(filiere_id, []).append(row)

    candid_by_filiere: dict[str, list[dict[str, Any]]] = {}
    for row in candidature:
        filiere_id = _safe_str(row.get("id_filiere"))
        candid_by_filiere.setdefault(filiere_id, []).append(row)

    metier_by_id = {_safe_str(r.get("id_metier")): r for r in metier}
    metiers_by_filiere: dict[str, list[dict[str, Any]]] = {}
    for row in mener:
        filiere_id = _safe_str(row.get("id_filiere"))
        metier_id = _safe_str(row.get("id_metier"))
        linked = metier_by_id.get(metier_id)
        if linked:
            metiers_by_filiere.setdefault(filiere_id, []).append(linked)

    schools: dict[str, dict] = {}
    transcripts: list[dict] = []

    for etab_id, e in etab_by_id.items():
        school_id = f"uni_{_normalize_token(etab_id)}"
        filieres = filieres_by_etab.get(etab_id, [])

        programs = sorted(
            {
                _normalize_token(f.get("domaine") or f.get("nom_filiere"))
                for f in filieres
                if _safe_str(f.get("domaine") or f.get("nom_filiere"))
            }
        )
        domain_labels = sorted(
            {
                _safe_str(f.get("domaine"))
                for f in filieres
                if _safe_str(f.get("domaine"))
            }
        )
        filiere_labels = sorted(
            {
                _safe_str(f.get("nom_filiere"))
                for f in filieres
                if _safe_str(f.get("nom_filiere"))
            }
        )

        tuition_values: list[int] = []
        salary_values: list[int] = []
        employability_values: list[float] = []
        has_concours = False

        for f in filieres:
            filiere_id = _safe_str(f.get("id_filiere"))
            for c in cout_by_filiere.get(filiere_id, []):
                tuition = _parse_int(c.get("frais_scolarite_annuel"), default=-1)
                if tuition >= 0:
                    tuition_values.append(tuition)

            for m in metiers_by_filiere.get(filiere_id, []):
                salary = _parse_int(m.get("salaire_debutant_maroc"), default=-1)
                if salary >= 0:
                    salary_values.append(salary)
                emp = _safe_str(m.get("employabilite")).lower()
                emp_score = {"tres_haute": 4.8, "haute": 4.4, "moyenne": 3.5, "faible": 2.6}.get(emp)
                if emp_score is not None:
                    employability_values.append(emp_score)

            concours_text = _safe_str(f.get("concours_ou_dossier")).lower()
            if "concours" in concours_text:
                has_concours = True

        school_type_raw = _safe_str(e.get("type")).lower()
        school_type = "public" if "public" in school_type_raw else "private"
        tuition_min = min(tuition_values) if tuition_values else (0 if school_type == "public" else 15000)
        tuition_max = max(tuition_values) if tuition_values else (12000 if school_type == "public" else 50000)

        international = _parse_bool(e.get("partenariats_internationaux")) or (
            "double" in _safe_str(e.get("accreditations")).lower()
        )

        schools[school_id] = {
            "school_id": school_id,
            "name": _safe_str(e.get("nom_etablissement")) or school_id,
            "country": "MA",
            "city": _safe_str(e.get("ville")),
            "type": school_type,
            "tuition_min_mad": tuition_min,
            "tuition_max_mad": tuition_max,
            "programs": programs or ["general"],
            "programs_tags": " | ".join(domain_labels),
            "filieres": " | ".join(filiere_labels),
            "admission_selectivity": "high" if has_concours else "medium",
            "employability_score": round(mean(employability_values), 2) if employability_values else 3.8,
            "salary_entry_min_mad": min(salary_values) if salary_values else 6000,
            "salary_entry_max_mad": max(salary_values) if salary_values else 12000,
            "international_double_degree": international,
        }

        school_name = schools[school_id]["name"]
        city = schools[school_id]["city"]
        status = school_type
        created_at = _to_iso_date(e.get("created_at"))
        source_url = _safe_str(e.get("site_web"))
        source_confidence = 0.8 if source_url else 0.6

        tuition_min_verified = min(tuition_values) if tuition_values else None
        tuition_max_verified = max(tuition_values) if tuition_values else None

        overview_text = (
            f"{school_name} est un établissement {status} situé à {city}." if city else f"{school_name} est un établissement {status}."
        )
        overview_metadata = _build_metadata(
            school_name=school_name,
            city=city,
            status=status,
            program="",
            domain="",
            tuition_min=tuition_min_verified,
            tuition_max=tuition_max_verified,
            admission_type="",
            source_url=source_url,
        )
        transcripts.append(
            _build_chunk(
                chunk_id=f"{school_id}_overview",
                school_id=school_id,
                chunk_type="overview",
                text=overview_text,
                language="fr",
                metadata=overview_metadata,
                created_at=created_at,
                source_url=source_url,
                source_confidence=source_confidence,
                program=(programs[0] if programs else "general"),
            )
        )

        for domain_label in domain_labels:
            if not domain_label:
                continue
            domain_metadata = _build_metadata(
                school_name=school_name,
                city=city,
                status=status,
                program="",
                domain=domain_label,
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_domain_{_normalize_token(domain_label)}",
                    school_id=school_id,
                    chunk_type="domain",
                    text=f"Domaine principal : {domain_label}.",
                    language="fr",
                    metadata=domain_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=(programs[0] if programs else "general"),
                )
            )

        if tuition_min_verified is not None or tuition_max_verified is not None:
            if tuition_min_verified is not None and tuition_max_verified is not None:
                tuition_text = (
                    f"Les frais de scolarité varient entre {tuition_min_verified} et {tuition_max_verified} MAD par an."
                )
            else:
                value = tuition_min_verified if tuition_min_verified is not None else tuition_max_verified
                tuition_text = f"Les frais de scolarité sont estimés à {value} MAD par an."
            tuition_metadata = _build_metadata(
                school_name=school_name,
                city=city,
                status=status,
                program="",
                domain="",
                tuition_min=tuition_min_verified,
                tuition_max=tuition_max_verified,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_tuition",
                    school_id=school_id,
                    chunk_type="tuition",
                    text=tuition_text,
                    language="fr",
                    metadata=tuition_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=(programs[0] if programs else "general"),
                )
            )

    for filiere_id, f in filiere_by_id.items():
        etab_id = _safe_str(f.get("id_etablissement"))
        school_id = f"uni_{_normalize_token(etab_id)}"
        school = schools.get(school_id)
        if not school:
            continue

        c_row = (cout_by_filiere.get(filiere_id) or [{}])[0]
        cand_row = (candid_by_filiere.get(filiere_id) or [{}])[0]

        program_label = _safe_str(f.get("nom_filiere"))
        domain_label = _safe_str(f.get("domaine"))
        admission_text = _safe_str(f.get("concours_ou_dossier"))

        language_raw = _safe_str(etab_by_id.get(etab_id, {}).get("langue_enseignement")).lower()
        if "english" in language_raw or "anglais" in language_raw or language_raw == "en":
            language = "en"
        elif "ar" in language_raw:
            language = "ar"
        elif "dar" in language_raw:
            language = "darija"
        else:
            language = "fr"

        created_at = _to_iso_date(cand_row.get("date_concours") or cand_row.get("date_limite"))
        source_url = _safe_str(etab_by_id.get(etab_id, {}).get("site_web"))
        source_confidence = 0.8 if source_url else 0.6

        if program_label:
            program_metadata = _build_metadata(
                school_name=school.get("name", ""),
                city=school.get("city", ""),
                status=school.get("type", ""),
                program=program_label,
                domain=domain_label,
                tuition_min=None,
                tuition_max=None,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"xlsx_{_normalize_token(filiere_id)}_program",
                    school_id=school_id,
                    chunk_type="program",
                    text=f"{school.get('name', '')} propose la formation {program_label}.",
                    language=language,
                    metadata=program_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=_normalize_token(program_label),
                )
            )

        if admission_text:
            admission_metadata = _build_metadata(
                school_name=school.get("name", ""),
                city=school.get("city", ""),
                status=school.get("type", ""),
                program=program_label,
                domain=domain_label,
                tuition_min=None,
                tuition_max=None,
                admission_type=admission_text,
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"xlsx_{_normalize_token(filiere_id)}_admission",
                    school_id=school_id,
                    chunk_type="admission",
                    text=f"Admission : {admission_text}.",
                    language=language,
                    metadata=admission_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=_normalize_token(program_label) if program_label else "general",
                )
            )

        tuition_value = _parse_int(c_row.get("frais_scolarite_annuel"), default=-1)
        if tuition_value >= 0:
            tuition_metadata = _build_metadata(
                school_name=school.get("name", ""),
                city=school.get("city", ""),
                status=school.get("type", ""),
                program=program_label,
                domain=domain_label,
                tuition_min=tuition_value,
                tuition_max=tuition_value,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"xlsx_{_normalize_token(filiere_id)}_tuition",
                    school_id=school_id,
                    chunk_type="tuition",
                    text=f"Les frais de scolarité sont estimés à {tuition_value} MAD par an.",
                    language=language,
                    metadata=tuition_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=_normalize_token(program_label) if program_label else "general",
                )
            )

    return schools, transcripts


def load_schools(csv_path: Path) -> dict[str, dict]:
    schools: dict[str, dict] = {}
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            school_id = row["school_id"].strip()
            raw_programs = _safe_str(row.get("programs"))
            programs = [p.strip().lower() for p in raw_programs.split("|") if p.strip()]
            row["programs"] = programs
            row["programs_tags"] = raw_programs
            row["filieres"] = raw_programs
            row["tuition_min_mad"] = int(row["tuition_min_mad"])
            row["tuition_max_mad"] = int(row["tuition_max_mad"])
            row["salary_entry_min_mad"] = int(row["salary_entry_min_mad"])
            row["salary_entry_max_mad"] = int(row["salary_entry_max_mad"])
            row["employability_score"] = float(row["employability_score"])
            row["international_double_degree"] = row["international_double_degree"].lower() == "true"
            schools[school_id] = row
    return schools


def load_transcripts(jsonl_path: Path) -> list[dict]:
    transcripts: list[dict] = []
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            item = json.loads(line)
            if len(item.get("text", "")) < 10:
                continue
            transcripts.append(item)
    return transcripts


def load_from_json_catalog(json_path: Path) -> tuple[dict[str, dict], list[dict]]:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        rows = payload.get("etablissements", [])
    elif isinstance(payload, list):
        rows = payload
    else:
        rows = []

    schools: dict[str, dict] = {}
    transcripts: list[dict] = []

    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue

        name = _safe_str(row.get("nom") or row.get("nom_complet") or row.get("name"))
        city = _safe_str(row.get("ville") or row.get("city"))
        if not name:
            continue

        school_id = f"uni_{_normalize_token(name)}_{_normalize_token(city)}"
        filieres = row.get("filieres", [])
        if isinstance(filieres, list):
            filieres_list = [_safe_str(v) for v in filieres if _safe_str(v)]
        else:
            filieres_list = _split_program_labels(filieres)

        formations_list = _split_program_labels(row.get("formations"))
        domain_label = _safe_str(row.get("domaine"))

        program_labels = []
        if domain_label:
            program_labels.append(domain_label)
        program_labels.extend(filieres_list)
        for label in formations_list:
            if label not in program_labels:
                program_labels.append(label)

        programs = [_normalize_token(v) for v in program_labels if _safe_str(v)]
        programs = sorted(set(programs)) or ["general"]
        unique_labels: list[str] = []
        for label in program_labels:
            if label and label not in unique_labels:
                unique_labels.append(label)

        status_text = _safe_str(row.get("statut") or row.get("type") or row.get("legal_status")).lower()
        school_type = "public" if "public" in status_text else "private"

        frais = row.get("frais") if isinstance(row.get("frais"), dict) else {}
        frais_min = _parse_int(frais.get("min_MAD_an"), default=-1)
        frais_max = _parse_int(frais.get("max_MAD_an"), default=-1)
        annuels = _parse_int(row.get("frais_annuels_mad"), default=-1)
        inscription = _parse_int(row.get("frais_inscription_mad"), default=-1)

        tuition_candidates = [v for v in [frais_min, frais_max, annuels, inscription] if v >= 0]
        tuition_min = min(tuition_candidates) if tuition_candidates else 0
        tuition_max = max(tuition_candidates) if tuition_candidates else 0

        schools[school_id] = {
            "school_id": school_id,
            "name": name,
            "country": "MA",
            "city": city,
            "type": school_type,
            "tuition_min_mad": tuition_min,
            "tuition_max_mad": tuition_max,
            "programs": programs,
            "programs_tags": " | ".join(unique_labels),
            "filieres": " | ".join(filieres_list),
            "admission_selectivity": "medium",
            "employability_score": 3.8,
            "salary_entry_min_mad": 6000,
            "salary_entry_max_mad": 12000,
            "international_double_degree": False,
        }

        created_at = _to_iso_date(row.get("date_collecte"))
        source_url = _safe_str(row.get("site_web") or row.get("website_url"))
        source_confidence = 0.75 if source_url else 0.6
        status = _safe_str(row.get("statut") or row.get("type")) or school_type

        overview_text = f"{name} est un établissement {status} situé à {city}." if city else f"{name} est un établissement {status}."
        if domain_label:
            overview_text += f" Domaines principaux : {domain_label}."

        overview_metadata = _build_metadata(
            school_name=name,
            city=city,
            status=status,
            program="",
            domain=domain_label,
            tuition_min=tuition_min if tuition_min > 0 else None,
            tuition_max=tuition_max if tuition_max > 0 else None,
            admission_type="",
            source_url=source_url,
        )
        transcripts.append(
            _build_chunk(
                chunk_id=f"{school_id}_overview",
                school_id=school_id,
                chunk_type="overview",
                text=overview_text,
                language="fr",
                metadata=overview_metadata,
                created_at=created_at,
                source_url=source_url,
                source_confidence=source_confidence,
                program=programs[0] if programs else "general",
            )
        )

        if domain_label:
            domain_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain_label,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_domain_{_normalize_token(domain_label)}",
                    school_id=school_id,
                    chunk_type="domain",
                    text=f"Domaine principal : {domain_label}.",
                    language="fr",
                    metadata=domain_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0] if programs else "general",
                )
            )

        for label in unique_labels:
            prog_label = _safe_str(label)
            if not prog_label:
                continue
            prog_token = _normalize_token(prog_label)
            program_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program=prog_label,
                domain=domain_label,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_program_{prog_token}",
                    school_id=school_id,
                    chunk_type="program",
                    text=f"{name} propose la formation {prog_label}.",
                    language="fr",
                    metadata=program_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=prog_token,
                )
            )

        admission_text = _safe_str(row.get("conditions_acces") or row.get("acces"))
        if admission_text:
            admission_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain_label,
                tuition_min=tuition_min if tuition_min > 0 else None,
                tuition_max=tuition_max if tuition_max > 0 else None,
                admission_type=admission_text,
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_admission",
                    school_id=school_id,
                    chunk_type="admission",
                    text=f"Admission : {admission_text}.",
                    language="fr",
                    metadata=admission_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0] if programs else "general",
                )
            )

        if tuition_min > 0 or tuition_max > 0:
            tuition_text = f"Les frais de scolarité varient entre {tuition_min} et {tuition_max} MAD par an."
            tuition_metadata = _build_metadata(
                school_name=name,
                city=city,
                status=status,
                program="",
                domain=domain_label,
                tuition_min=tuition_min,
                tuition_max=tuition_max,
                admission_type="",
                source_url=source_url,
            )
            transcripts.append(
                _build_chunk(
                    chunk_id=f"{school_id}_tuition",
                    school_id=school_id,
                    chunk_type="tuition",
                    text=tuition_text,
                    language="fr",
                    metadata=tuition_metadata,
                    created_at=created_at,
                    source_url=source_url,
                    source_confidence=source_confidence,
                    program=programs[0] if programs else "general",
                )
            )

    return schools, transcripts


def load_policy(policy_path: Path) -> dict:
    with policy_path.open("r", encoding="utf-8") as f:
        return {"raw_text": f.read()}


def load_bundle(root_dir: Path) -> DataBundle:
    strict_supabase = os.getenv("SUPABASE_STRICT_MODE", "1").strip().lower() in {"1", "true", "yes", "on"}
    supabase_url = (os.getenv("SUPABASE_URL") or "").strip()
    supabase_key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY")
        or os.getenv("SUPABASE_API_KEY")
        or os.getenv("SUPABASE_ANON_KEY")
        or ""
    ).strip()
    if strict_supabase and (not supabase_url or not supabase_key):
        strict_supabase = False

    if not strict_supabase:
        json_catalog_path = root_dir / "etablissements_maroc_complet.json"
        xlsx_candidates = [
            root_dir / "BDD_MCD_Universites.xlsx",
            root_dir / "etablissements_maroc_complet.xlsx",
        ]
        xlsx_path = next((p for p in xlsx_candidates if p.exists()), None)

        if json_catalog_path.exists():
            schools, transcripts = load_from_json_catalog(json_catalog_path)
        elif xlsx_path is not None:
            schools, transcripts = load_from_excel_mcd(xlsx_path)
        else:
            schools = {}
            transcripts = []

        policy = load_policy(root_dir / "config" / "policy_rules.yaml")
        return DataBundle(schools=schools, transcripts=transcripts, policy=policy, source="local_fallback")

    try:
        supabase_limit = _parse_int(os.getenv("SUPABASE_SCHOOLS_LIMIT", "500"), default=500)
        sb_schools, sb_transcripts = load_from_supabase_schools(limit=max(1, min(5000, supabase_limit)))
        if sb_schools:
            policy = load_policy(root_dir / "config" / "policy_rules.yaml")
            return DataBundle(schools=sb_schools, transcripts=sb_transcripts, policy=policy, source="supabase_schools")
    except Exception as exc:
        if strict_supabase:
            raise RuntimeError(f"Supabase schools load failed in strict mode: {exc}") from exc

    raise RuntimeError(
        "Supabase strict mode is enabled, but no schools were loaded from DB. "
        "Set SUPABASE_URL and SUPABASE_ANON_KEY, or switch to local mode with SUPABASE_STRICT_MODE=0."
    )
